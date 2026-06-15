import streamlit as st
import requests
import os
import re
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# Configuration - Support both Streamlit Cloud (secrets) and local (env vars)
def get_backend_url():
    """Get backend URL from Streamlit secrets or environment variable"""
    try:
        # Try Streamlit secrets first (for Streamlit Cloud deployment)
        return st.secrets["BACKEND_URL"]
    except (KeyError, FileNotFoundError):
        # Fallback to environment variable (for local/docker deployment)
        return os.getenv("BACKEND_URL", "http://localhost:8000")

BACKEND_URL = get_backend_url()

# Page configuration
st.set_page_config(
    page_title="Privacy Intake Chatbot",
    page_icon="🔒",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS for dark mode fix and better styling
st.markdown("""
<style>
    /* Fix dark mode styling */
    .stChatMessage {
        background-color: transparent !important;
    }
    
    /* Better input styling */
    .stChatInputContainer {
        border-top: 1px solid rgba(49, 51, 63, 0.2);
        padding-top: 1rem;
    }
    
    /* Session list styling */
    .session-item {
        padding: 0.5rem;
        border-radius: 0.5rem;
        margin-bottom: 0.5rem;
        cursor: pointer;
        transition: background-color 0.2s;
    }
    
    .session-item:hover {
        background-color: rgba(151, 166, 195, 0.15);
    }
    
    .session-item.active {
        background-color: rgba(28, 131, 225, 0.15);
        border-left: 3px solid #1c83e1;
    }
    
    /* Welcome message */
    .welcome-box {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 2rem;
        border-radius: 1rem;
        color: white;
        margin-bottom: 1.5rem;
    }
    
    .welcome-box h4 {
        color: white;
        margin-bottom: 0.5rem;
    }
</style>
""", unsafe_allow_html=True)


def parse_markdown_to_excel(markdown_content: str) -> bytes:
    """Convert markdown analysis result to Excel file"""
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    section_font = Font(bold=True, size=12)
    
    ws_summary = wb.active
    ws_summary.title = "Tóm Tắt"
    
    sections = re.split(r'\n## ', markdown_content)
    
    current_row = 1
    
    ws_summary.cell(row=current_row, column=1, value="KẾT QUẢ PHÂN TÍCH CASE - DATA PRIVACY")
    ws_summary.cell(row=current_row, column=1).font = Font(bold=True, size=16)
    ws_summary.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=4)
    current_row += 2
    
    classification_data = []
    checklist_data = []
    summary_text = ""
    
    for section in sections:
        if section.startswith("A.") or "Bảng Phân Loại" in section:
            table_match = re.findall(r'\|(.+?)\|(.+?)\|', section)
            for row in table_match:
                if '---' not in row[0] and 'Câu hỏi' not in row[0]:
                    classification_data.append({
                        'Câu hỏi': row[0].strip(),
                        'Kết quả sơ bộ': row[1].strip()
                    })
        
        elif section.startswith("D.") or "Checklist" in section:
            table_match = re.findall(r'\|(.+?)\|(.+?)\|', section)
            for row in table_match:
                if '---' not in row[0] and 'Nhóm hồ sơ' not in row[0]:
                    checklist_data.append({
                        'Nhóm hồ sơ': row[0].strip(),
                        'Cần chuẩn bị': row[1].strip()
                    })
        
        elif section.startswith("G.") or "Tóm Tắt" in section:
            summary_match = re.search(r'\*\*Tóm tắt request gửi Data Privacy:\*\*\s*(.*?)(?:\n---|\n##|$)', section, re.DOTALL)
            if summary_match:
                summary_text = summary_match.group(1).strip()
    
    ws_summary.cell(row=current_row, column=1, value="BẢNG PHÂN LOẠI SƠ BỘ")
    ws_summary.cell(row=current_row, column=1).font = section_font
    current_row += 1
    
    if classification_data:
        ws_summary.cell(row=current_row, column=1, value="Câu hỏi")
        ws_summary.cell(row=current_row, column=2, value="Kết quả sơ bộ")
        ws_summary.cell(row=current_row, column=1).font = header_font
        ws_summary.cell(row=current_row, column=1).fill = header_fill
        ws_summary.cell(row=current_row, column=2).font = header_font
        ws_summary.cell(row=current_row, column=2).fill = header_fill
        current_row += 1
        
        for item in classification_data:
            ws_summary.cell(row=current_row, column=1, value=item['Câu hỏi'])
            ws_summary.cell(row=current_row, column=2, value=item['Kết quả sơ bộ'])
            current_row += 1
    
    current_row += 2
    
    ws_summary.cell(row=current_row, column=1, value="CHECKLIST HỒ SƠ CẦN CHUẨN BỊ")
    ws_summary.cell(row=current_row, column=1).font = section_font
    current_row += 1
    
    if checklist_data:
        ws_summary.cell(row=current_row, column=1, value="Nhóm hồ sơ")
        ws_summary.cell(row=current_row, column=2, value="Cần chuẩn bị")
        ws_summary.cell(row=current_row, column=1).font = header_font
        ws_summary.cell(row=current_row, column=1).fill = header_fill
        ws_summary.cell(row=current_row, column=2).font = header_font
        ws_summary.cell(row=current_row, column=2).fill = header_fill
        current_row += 1
        
        for item in checklist_data:
            ws_summary.cell(row=current_row, column=1, value=item['Nhóm hồ sơ'])
            ws_summary.cell(row=current_row, column=2, value=item['Cần chuẩn bị'])
            current_row += 1
    
    ws_summary.column_dimensions['A'].width = 45
    ws_summary.column_dimensions['B'].width = 50
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()


# Sample cases
SAMPLE_CASES = {
    "🟢 No Personal Data": """Team tôi muốn chia sẻ báo cáo tổng hợp số lượng giao dịch theo tháng cho đối tác, không có user ID, số điện thoại, email hoặc thông tin định danh khách hàng. Báo cáo chỉ bao gồm tổng số giao dịch, giá trị trung bình và phân bố theo khu vực. Đối tác là công ty phân tích thị trường tại Việt Nam.""",
    
    "🟡 Domestic Sharing": """Team tôi muốn chia sẻ họ tên, số điện thoại và lịch sử giao dịch của khách hàng cho vendor tại Việt Nam để chăm sóc khách hàng. Vendor sẽ gọi điện và hỗ trợ khách hàng về các giao dịch của họ. Dữ liệu bao gồm: tên, số điện thoại, email, user ID, lịch sử giao dịch (ngày, số tiền, loại giao dịch) trong 6 tháng gần nhất. Vendor có khoảng 50 nhân viên chăm sóc khách hàng tại văn phòng Hà Nội. Chúng tôi chưa có DPA với vendor này.""",
    
    "🔴 ANT Singapore": """Team em muốn hợp tác với đối tác ANT Singapore để tư vấn credit scoring. Dữ liệu dự kiến chia sẻ gồm user_id_hash, transaction_count, transaction_amount, BNPL_payment_amt, device_id. Dữ liệu gửi qua API hằng ngày. Đối tác lưu 12 tháng để tư vấn scoring. Chưa rõ có DPA chưa. Không biết cần chuẩn bị gì trước khi gửi Data Privacy review?"""
}


# API Functions
def get_sessions():
    """Get all chat sessions"""
    try:
        response = requests.get(f"{BACKEND_URL}/sessions", timeout=10)
        if response.status_code == 200:
            return response.json().get("sessions", [])
    except:
        pass
    return []


def create_session(name: str):
    """Create a new session"""
    try:
        response = requests.post(
            f"{BACKEND_URL}/sessions",
            json={"name": name},
            timeout=10
        )
        if response.status_code == 200:
            return response.json().get("session")
    except:
        pass
    return None


def get_session(session_id: str):
    """Get a specific session"""
    try:
        response = requests.get(f"{BACKEND_URL}/sessions/{session_id}", timeout=10)
        if response.status_code == 200:
            return response.json()
    except:
        pass
    return None


def update_session(session_id: str, messages: list):
    """Update session messages"""
    try:
        response = requests.put(
            f"{BACKEND_URL}/sessions/{session_id}",
            json={"messages": messages},
            timeout=10
        )
        if response.status_code == 200:
            return True
    except:
        pass
    return False


def delete_session(session_id: str):
    """Delete a session"""
    try:
        response = requests.delete(f"{BACKEND_URL}/sessions/{session_id}", timeout=10)
        return response.status_code == 200
    except:
        return False


def get_settings():
    """Get current settings"""
    try:
        response = requests.get(f"{BACKEND_URL}/settings", timeout=10)
        if response.status_code == 200:
            return response.json()
    except:
        pass
    return {"model": "qwen/qwen3-5-27b", "custom_instructions": "", "available_models": []}


def send_message(user_message: str, conversation_history: list):
    """Send message to backend and get response"""
    try:
        response = requests.post(
            f"{BACKEND_URL}/chat",
            json={
                "message": user_message,
                "conversation_history": conversation_history
            },
            timeout=120
        )
        
        if response.status_code == 200:
            result = response.json()
            return result.get("answer", "Không có phản hồi từ agent")
        else:
            return f"❌ Lỗi: Backend trả về status code {response.status_code}"
            
    except requests.exceptions.ConnectionError:
        return "❌ Không thể kết nối backend. Vui lòng đảm bảo backend đang chạy."
    except requests.exceptions.Timeout:
        return "⏱️ Request timeout. Vui lòng thử lại."
    except Exception as e:
        return f"❌ Đã xảy ra lỗi: {str(e)}"


# Initialize session state
if "current_session_id" not in st.session_state:
    st.session_state.current_session_id = None

if "messages" not in st.session_state:
    st.session_state.messages = []

if "pending_sample" not in st.session_state:
    st.session_state.pending_sample = None

if "settings" not in st.session_state:
    st.session_state.settings = get_settings()

if "just_created_session" not in st.session_state:
    st.session_state.just_created_session = False


# Sidebar
with st.sidebar:
    st.title("🔒 Privacy Chatbot")
    
    # Session management
    st.subheader("💬 Chat Sessions")
    
    # New session button
    col1, col2 = st.columns([3, 1])
    with col1:
        if st.button("➕ New Chat", use_container_width=True, key="new_chat_btn"):
            new_session = create_session(f"Chat {datetime.now().strftime('%Y-%m-%d %H:%M')}")
            if new_session:
                st.session_state.current_session_id = new_session["id"]
                st.session_state.messages = []
                st.toast("✅ Chat mới đã được tạo!", icon="✅")
                st.rerun()
            else:
                st.error("❌ Không thể tạo chat mới")
    
    with col2:
        if st.button("🔄", use_container_width=True, help="Refresh danh sách"):
            st.rerun()
    
    # List sessions
    sessions = get_sessions()
    
    # Show active session info
    if st.session_state.current_session_id:
        current_name = "Unknown"
        for s in sessions:
            if s["id"] == st.session_state.current_session_id:
                current_name = s["name"]
                break
        st.info(f"📍 **Active:** {current_name}")
    
    if sessions:
        st.markdown("**Saved Chats:**")
        for session in sessions:
            is_active = session["id"] == st.session_state.current_session_id
            
            col1, col2 = st.columns([4, 1])
            
            with col1:
                # Button label with message count
                msg_count = session.get("message_count", 0)
                label = f"{'📍' if is_active else '💬'} {session['name']} ({msg_count})"
                
                if st.button(
                    label,
                    key=f"session_{session['id']}",
                    use_container_width=True,
                    type="primary" if is_active else "secondary",
                    disabled=is_active  # Disable if already active
                ):
                    # Load session
                    loaded_session = get_session(session["id"])
                    if loaded_session:
                        st.session_state.current_session_id = session["id"]
                        st.session_state.messages = loaded_session.get("messages", [])
                        st.toast(f"📂 Đã load: {session['name']}", icon="📂")
                        st.rerun()
            
            with col2:
                if st.button("🗑️", key=f"delete_{session['id']}", use_container_width=True, help="Xóa chat"):
                    if delete_session(session["id"]):
                        if st.session_state.current_session_id == session["id"]:
                            st.session_state.current_session_id = None
                            st.session_state.messages = []
                        st.toast("🗑️ Đã xóa chat", icon="🗑️")
                        st.rerun()
    else:
        st.info("💡 Chưa có chat nào. Click 'New Chat' để bắt đầu!")
    
    st.divider()
    
    # Demo cases
    st.subheader("📋 Demo Cases")
    for case_name, case_text in SAMPLE_CASES.items():
        if st.button(case_name, key=f"sample_{case_name}", use_container_width=True):
            st.session_state.pending_sample = case_text
    
    st.divider()
    
    # Export section
    if st.session_state.messages:
        st.subheader("📥 Export")
        
        last_assistant_msg = None
        for msg in reversed(st.session_state.messages):
            if msg["role"] == "assistant":
                last_assistant_msg = msg["content"]
                break
        
        if last_assistant_msg:
            col1, col2 = st.columns(2)
            with col1:
                excel_data = parse_markdown_to_excel(last_assistant_msg)
                st.download_button(
                    label="📥 Excel",
                    data=excel_data,
                    file_name="privacy_analysis.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            with col2:
                st.download_button(
                    label="📄 MD",
                    data=last_assistant_msg,
                    file_name="privacy_analysis.md",
                    mime="text/markdown",
                    use_container_width=True
                )
    
    st.divider()
    
    # Admin Panel link
    if st.button("⚙️ Admin Panel", use_container_width=True, help="Edit prompts, knowledge, skills"):
        st.switch_page("pages/1_⚙️_Admin_Panel.py")
    
    st.divider()
    
    st.caption(f"**Model:** {st.session_state.settings.get('model', 'N/A')}")
    st.caption("**Version:** 2.0.0")


# Main chat area
st.title("Privacy Intake Chatbot")
st.caption("Chatbot hỗ trợ sàng lọc yêu cầu chia sẻ dữ liệu trước khi gửi Data Privacy review")

# Display welcome message or chat history
if not st.session_state.messages:
    # Show session info if exists
    if st.session_state.current_session_id:
        st.success("✅ Chat mới đã sẵn sàng! Bắt đầu nhập câu hỏi ở dưới.", icon="✅")
    
    st.markdown("""
    <div class="welcome-box">
        <h4>👋 Xin chào! Tôi là Privacy Intake Chatbot</h4>
        <p>Tôi sẽ giúp bạn chuẩn bị hồ sơ trước khi gửi Data Privacy review.</p>
        <p><strong>Hãy mô tả case của bạn:</strong></p>
        <ul>
            <li>Bạn muốn chia sẻ dữ liệu gì?</li>
            <li>Cho đối tác nào?</li>
            <li>Mục đích là gì?</li>
        </ul>
        <p><em>💡 Tip: Chọn Demo Case từ sidebar để test nhanh</em></p>
    </div>
    """, unsafe_allow_html=True)
else:
    # Display chat messages
    for message in st.session_state.messages:
        with st.chat_message(message["role"]):
            st.markdown(message["content"])

# Handle pending sample case
if st.session_state.pending_sample:
    sample_text = st.session_state.pending_sample
    st.session_state.pending_sample = None
    
    # Add user message
    st.session_state.messages.append({"role": "user", "content": sample_text})
    
    # Save to current session
    if st.session_state.current_session_id:
        update_session(st.session_state.current_session_id, st.session_state.messages)
    
    # Get response
    with st.spinner("🤖 Đang phân tích..."):
        history = [{"role": m["role"], "content": m["content"]} for m in st.session_state.messages[:-1]]
        response = send_message(sample_text, history)
    
    # Add assistant response
    st.session_state.messages.append({"role": "assistant", "content": response})
    
    # Save to current session
    if st.session_state.current_session_id:
        update_session(st.session_state.current_session_id, st.session_state.messages)
    
    st.rerun()

# Chat input
if prompt := st.chat_input("Nhập mô tả case của bạn..."):
    # Create new session if none exists
    if not st.session_state.current_session_id:
        new_session = create_session(f"Chat {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        if new_session:
            st.session_state.current_session_id = new_session["id"]
    
    # Add user message
    st.session_state.messages.append({"role": "user", "content": prompt})
    
    # Display user message
    with st.chat_message("user"):
        st.markdown(prompt)
    
    # Get and display assistant response
    with st.chat_message("assistant"):
        with st.spinner("🤖 Đang phân tích..."):
            history = [{"role": m["role"], "content": m["content"]} for m in st.session_state.messages[:-1]]
            response = send_message(prompt, history)
        st.markdown(response)
    
    # Add assistant response
    st.session_state.messages.append({"role": "assistant", "content": response})
    
    # Save to current session
    if st.session_state.current_session_id:
        update_session(st.session_state.current_session_id, st.session_state.messages)

# Footer
st.divider()
st.markdown("""
<div style='text-align: center; color: #666; font-size: 0.9em;'>
    <p><strong>⚠️ Lưu ý:</strong> Đây là phân loại sơ bộ. Kết luận cuối cùng cần Data Privacy team xác nhận.</p>
    <p>Privacy Intake Chatbot v2.0 | Built with ❤️ for AI Competition</p>
</div>
""", unsafe_allow_html=True)
